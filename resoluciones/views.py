from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.urls import reverse_lazy
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

from .models import Resolucion, DetalleResolucion, ResponsableObra, Zonificacion, LicenciaUso
from .forms import CustomUserCreationForm, ResolucionForm, DetalleResolucionForm, ResolucionFilterForm, ExportarExcelForm


def registro(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registro exitoso. Contacte al administrador para asignar permisos.')
            return redirect('dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/registro.html', {'form': form})


@login_required
def dashboard(request):
    user_groups = [group.name for group in request.user.groups.all()]
    
    if 'Obras' in user_groups:
        return render(request, 'resoluciones/dashboard_obras.html')
    elif 'Fiscalización' in user_groups:
        return redirect('lista_resoluciones')
    else:
        messages.warning(request, 'No tiene permisos asignados. Contacte al administrador.')
        return render(request, 'resoluciones/dashboard_obras.html')


class ResolucionListView(LoginRequiredMixin, ListView):
    model = Resolucion
    template_name = 'resoluciones/lista_resoluciones.html'
    context_object_name = 'resoluciones'
    paginate_by = 10

    def get_queryset(self):
        queryset = Resolucion.objects.all()
        
        # Aplicar filtros
        form = ResolucionFilterForm(self.request.GET)
        if form.is_valid():
            if form.cleaned_data['num_resolucion']:
                queryset = queryset.filter(
                    num_resolucion__icontains=form.cleaned_data['num_resolucion']
                )
            if form.cleaned_data['num_expediente']:
                queryset = queryset.filter(
                    num_expediente__icontains=form.cleaned_data['num_expediente']
                )
            if form.cleaned_data['administrado']:
                queryset = queryset.filter(
                    administrado__icontains=form.cleaned_data['administrado']
                )
            if form.cleaned_data['nombre_licencia']:
                queryset = queryset.filter(
                    nombre_licencia=form.cleaned_data['nombre_licencia']
                )
        
        return queryset.order_by('-fecha_emision')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ResolucionFilterForm(self.request.GET)
        context['user_groups'] = [group.name for group in self.request.user.groups.all()]
        return context


class ResolucionDetailView(LoginRequiredMixin, DetailView):
    model = Resolucion
    template_name = 'resoluciones/detalle_resolucion.html'
    context_object_name = 'resolucion'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_groups'] = [group.name for group in self.request.user.groups.all()]
        try:
            context['detalle'] = self.object.detalleresolucion
        except DetalleResolucion.DoesNotExist:
            context['detalle'] = None
        return context


@login_required
def crear_resolucion(request):
    user_groups = [group.name for group in request.user.groups.all()]
    
    # Solo usuarios de Obras pueden crear resoluciones
    if 'Obras' not in user_groups:
        messages.error(request, 'No tiene permisos para crear resoluciones.')
        return redirect('dashboard')

    if request.method == 'POST':
        resolucion_form = ResolucionForm(request.POST)
        detalle_form = DetalleResolucionForm(request.POST)
        
        if resolucion_form.is_valid() and detalle_form.is_valid():
            try:
                with transaction.atomic():
                    # Crear resolución
                    resolucion = resolucion_form.save(commit=False)
                    resolucion.usuario = request.user
                    resolucion.save()
                    
                    # Crear detalle
                    detalle = detalle_form.save(commit=False)
                    detalle.resolucion = resolucion
                    detalle.save()
                    
                    messages.success(request, 'Resolución creada exitosamente.')
                    return redirect('detalle_resolucion', pk=resolucion.pk)
            except Exception as e:
                messages.error(request, f'Error al crear la resolución: {str(e)}')
    else:
        resolucion_form = ResolucionForm()
        detalle_form = DetalleResolucionForm()

    context = {
        'resolucion_form': resolucion_form,
        'detalle_form': detalle_form,
        'title': 'Nueva Resolución'
    }
    return render(request, 'resoluciones/form_resolucion.html', context)


@login_required
def editar_resolucion(request, pk):
    resolucion = get_object_or_404(Resolucion, pk=pk)
    user_groups = [group.name for group in request.user.groups.all()]
    
    # Solo usuarios de Obras pueden editar cualquier resolución
    if 'Obras' not in user_groups:
        messages.error(request, 'No tiene permisos para editar esta resolución.')
        return redirect('detalle_resolucion', pk=pk)

    try:
        detalle = resolucion.detalleresolucion
    except DetalleResolucion.DoesNotExist:
        detalle = None

    if request.method == 'POST':
        resolucion_form = ResolucionForm(request.POST, instance=resolucion)
        detalle_form = DetalleResolucionForm(request.POST, instance=detalle) if detalle else DetalleResolucionForm(request.POST)
        
        if resolucion_form.is_valid() and detalle_form.is_valid():
            try:
                with transaction.atomic():
                    resolucion_form.save()
                    
                    if detalle:
                        detalle_form.save()
                    else:
                        new_detalle = detalle_form.save(commit=False)
                        new_detalle.resolucion = resolucion
                        new_detalle.save()
                    
                    messages.success(request, 'Resolución actualizada exitosamente.')
                    return redirect('detalle_resolucion', pk=resolucion.pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar la resolución: {str(e)}')
    else:
        resolucion_form = ResolucionForm(instance=resolucion)
        detalle_form = DetalleResolucionForm(instance=detalle) if detalle else DetalleResolucionForm()

    context = {
        'resolucion_form': resolucion_form,
        'detalle_form': detalle_form,
        'resolucion': resolucion,
        'title': 'Editar Resolución'
    }
    return render(request, 'resoluciones/form_resolucion.html', context)


@login_required
def mis_resoluciones(request):
    user_groups = [group.name for group in request.user.groups.all()]
    
    if 'Obras' not in user_groups:
        messages.error(request, 'No tiene permisos para acceder a esta página.')
        return redirect('dashboard')

     # 1. Cargar el formulario con GET
    filter_form = ResolucionFilterForm(request.GET or None)

    # 2. Filtro base
    resoluciones = Resolucion.objects.filter(usuario=request.user)
    
    # 3. Aplicar filtros si válidos
    if filter_form.is_valid():
            cd = filter_form.cleaned_data
            if cd.get('num_resolucion'):
                resoluciones = resoluciones.filter(num_resolucion__icontains=cd['num_resolucion'])
            if cd.get('num_expediente'):
                resoluciones = resoluciones.filter(num_expediente__icontains=cd['num_expediente'])
            if cd.get('administrado'):
                resoluciones = resoluciones.filter(administrado__icontains=cd['administrado'])
            if cd.get('nombre_licencia'):
                resoluciones = resoluciones.filter(nombre_licencia=cd['nombre_licencia'])

    # Orden y paginación
    resoluciones = resoluciones.order_by('-fecha_emision')
    paginator = Paginator(resoluciones, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'resoluciones': page_obj,
        'filter_form': filter_form,
        'user_groups': user_groups,
        'title': 'Mis Resoluciones'
    }
    return render(request, 'resoluciones/lista_resoluciones.html', context)


@login_required
def exportar_excel_form(request):
    """Vista para mostrar el formulario de exportación a Excel"""
    user_groups = [group.name for group in request.user.groups.all()]
    
    # Solo usuarios del grupo Obras pueden acceder
    if 'Obras' not in user_groups:
        messages.error(request, 'No tiene permisos para acceder a esta funcionalidad.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ExportarExcelForm(request.POST)
        if form.is_valid():
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            
            # Filtrar resoluciones por rango de fechas
            resoluciones = Resolucion.objects.filter(
                fecha_emision__gte=fecha_inicio,
                fecha_emision__lte=fecha_fin
            ).order_by('fecha_emision')
            
            if not resoluciones.exists():
                messages.warning(request, f'No se encontraron resoluciones en el rango de fechas del {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}.')
                return render(request, 'resoluciones/exportar_excel.html', {'form': form})
            
            # Generar archivo Excel
            return generar_excel_resoluciones(resoluciones)
    else:
        form = ExportarExcelForm()
    
    context = {
        'form': form,
        'title': 'Exportar Resoluciones a Excel'
    }
    return render(request, 'resoluciones/exportar_excel.html', context)


def generar_excel_resoluciones(resoluciones):
    """Función para generar el archivo Excel con las resoluciones"""
    # Crear un nuevo workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resoluciones de Obras"
    
    # Definir encabezados
    encabezados = ['N°', 'Serie Documental', 'Fecha', 'Asunto', 'Folios', 'Observaciones']
    
    # Escribir encabezados en la primera fila
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=encabezado)
        cell.font = Font(bold=True)
    
    # Escribir datos de las resoluciones
    for row, resolucion in enumerate(resoluciones, 2):
        ws.cell(row=row, column=1, value=row-1)  # N° correlativo
        ws.cell(row=row, column=2, value=f"RESOLUCION N° {resolucion.num_resolucion}-MPL-GDTI-SGDTyC-UFOPCUyL ") # Serie Documental
        ws.cell(row=row, column=3, value=resolucion.fecha_emision.strftime('%d/%m/%Y'))  # Fecha
        ws.cell(row=row, column=4, value=resolucion.administrado)  # Asunto
        ws.cell(row=row, column=5, value='1')  # Folios (siempre 1)
        ws.cell(row=row, column=6, value='-')  # Observaciones (siempre -)
    
    # Ajustar ancho de columnas
    column_widths = [5, 20, 12, 30, 8, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="resoluciones_obras.xlsx"'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    return response
